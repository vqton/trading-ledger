from datetime import date, datetime
from decimal import Decimal
from typing import Dict, List, Optional
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from services.ledger_service import LedgerService
from services.financial_report_service import FinancialReportService


class ExcelExporter:
    """Excel export utility for financial reports."""

    @staticmethod
    def _create_header_style():
        """Create header style."""
        return {
            "font": Font(bold=True, size=12),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "font_color": "FFFFFF",
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

    @staticmethod
    def _create_cell_style():
        """Create cell style."""
        return {
            "alignment": Alignment(horizontal="right", vertical="center"),
            "number_format": "#,##0",
        }

    @staticmethod
    def _create_title_style():
        """Create title style."""
        return {
            "font": Font(bold=True, size=14),
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

    @staticmethod
    def export_trial_balance(end_date: date) -> bytes:
        """Export Trial Balance to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bang Can Do Tai Khoan"

        header_style = ExcelExporter._create_header_style()
        title_style = ExcelExporter._create_title_style()
        cell_style = ExcelExporter._create_cell_style()

        ws.merge_cells("A1:E1")
        cell = ws["A1"]
        cell.value = "BẢNG CÂN ĐỐI TÀI KHOẢN"
        cell.font = title_style["font"]
        cell.alignment = title_style["alignment"]

        ws.merge_cells("A2:E2")
        cell = ws["A2"]
        cell.value = f"Ngày lập: {end_date.strftime('%d/%m/%Y')}"
        cell.alignment = Alignment(horizontal="center")

        headers = ["Mã TK", "Tên tài khoản", "Loại", "Số dư Nợ", "Số dư Có"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            for key, value in header_style.items():
                setattr(cell, key, value)

        data = LedgerService.get_trial_balance(end_date)

        for row_idx, acc in enumerate(data["accounts"], 5):
            ws.cell(row=row_idx, column=1, value=acc["account_code"])
            ws.cell(row=row_idx, column=2, value=acc["account_name"])
            ws.cell(row=row_idx, column=3, value=acc["account_type"])

            debit_cell = ws.cell(row=row_idx, column=4)
            debit_cell.value = float(acc["debit_balance"]) if acc["debit_balance"] > 0 else None
            if acc["debit_balance"] > 0:
                debit_cell.number_format = "#,##0"

            credit_cell = ws.cell(row=row_idx, column=5)
            credit_cell.value = float(acc["credit_balance"]) if acc["credit_balance"] > 0 else None
            if acc["credit_balance"] > 0:
                credit_cell.number_format = "#,##0"

        last_row = len(data["accounts"]) + 5
        ws.cell(row=last_row, column=3, value="Tổng cộng:")
        ws.cell(row=last_row, column=3).font = Font(bold=True)
        ws.cell(row=last_row, column=4, value=float(data["total_debit"]))
        ws.cell(row=last_row, column=4).number_format = "#,##0"
        ws.cell(row=last_row, column=4).font = Font(bold=True)
        ws.cell(row=last_row, column=5, value=float(data["total_credit"]))
        ws.cell(row=last_row, column=5).number_format = "#,##0"
        ws.cell(row=last_row, column=5).font = Font(bold=True)

        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_balance_sheet(end_date: date) -> bytes:
        """Export Balance Sheet to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bang Can Do Ke Toan"

        title_style = ExcelExporter._create_title_style()
        header_style = ExcelExporter._create_header_style()
        cell_style = ExcelExporter._create_cell_style()

        ws.merge_cells("A1:C1")
        cell = ws["A1"]
        cell.value = "BẢNG CÂN ĐỐI KẾ TOÁN"
        cell.font = title_style["font"]
        cell.alignment = title_style["alignment"]

        ws.merge_cells("A2:C2")
        cell = ws["A2"]
        cell.value = f"Tại ngày {end_date.strftime('%d/%m/%Y')}"
        cell.alignment = Alignment(horizontal="center")

        data = FinancialReportService.get_balance_sheet(end_date)

        row = 4

        ws.cell(row=row, column=1, value="A. TÀI SẢN")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        for acc in data["assets"]:
            ws.cell(row=row, column=1, value=acc["account_code"])
            ws.cell(row=row, column=2, value=acc["account_name"])
            cell = ws.cell(row=row, column=3)
            cell.value = float(acc["balance"]) if acc["balance"] != 0 else None
            if acc["balance"] != 0:
                cell.number_format = "#,##0"
            row += 1

        ws.cell(row=row, column=2, value="Tổng Tài sản (A)")
        ws.cell(row=row, column=2).font = Font(bold=True)
        cell = ws.cell(row=row, column=3)
        cell.value = float(data["total_assets"])
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        row += 2

        ws.cell(row=row, column=1, value="B. NỢ PHẢI TRẢ")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        for acc in data["liabilities"]:
            ws.cell(row=row, column=1, value=acc["account_code"])
            ws.cell(row=row, column=2, value=acc["account_name"])
            cell = ws.cell(row=row, column=3)
            cell.value = float(acc["balance"]) if acc["balance"] != 0 else None
            if acc["balance"] != 0:
                cell.number_format = "#,##0"
            row += 1

        ws.cell(row=row, column=2, value="Tổng Nợ phải trả (B)")
        ws.cell(row=row, column=2).font = Font(bold=True)
        cell = ws.cell(row=row, column=3)
        cell.value = float(data["total_liabilities"])
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        row += 2

        ws.cell(row=row, column=1, value="C. VỐN CHỦ SỞ HỮU")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        for acc in data["equity"]:
            ws.cell(row=row, column=1, value=acc["account_code"])
            ws.cell(row=row, column=2, value=acc["account_name"])
            cell = ws.cell(row=row, column=3)
            cell.value = float(acc["balance"]) if acc["balance"] != 0 else None
            if acc["balance"] != 0:
                cell.number_format = "#,##0"
            row += 1

        ws.cell(row=row, column=2, value="Lợi nhuận sau thuế")
        cell = ws.cell(row=row, column=3)
        cell.value = float(data["retained_earnings"])
        cell.number_format = "#,##0"
        row += 1

        ws.cell(row=row, column=2, value="Tổng Vốn chủ sở hữu (C)")
        ws.cell(row=row, column=2).font = Font(bold=True)
        cell = ws.cell(row=row, column=3)
        cell.value = float(data["total_equity"] + data["retained_earnings"])
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)

        for col in range(1, 4):
            ws.column_dimensions[chr(64 + col)].width = 25

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_income_statement(start_date: date, end_date: date) -> bytes:
        """Export Income Statement to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bao Cao Ket Qua KD"

        title_style = ExcelExporter._create_title_style()
        cell_style = ExcelExporter._create_cell_style()

        ws.merge_cells("A1:C1")
        cell = ws["A1"]
        cell.value = "BÁO CÁO KẾT QUẢ KINH DOANH"
        cell.font = title_style["font"]
        cell.alignment = title_style["alignment"]

        ws.merge_cells("A2:C2")
        cell = ws["A2"]
        cell.value = f"Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}"
        cell.alignment = Alignment(horizontal="center")

        data = FinancialReportService.get_income_statement(start_date, end_date)

        row = 4

        ws.cell(row=row, column=1, value="1. DOANH THU")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for acc in data["revenue"]:
            ws.cell(row=row, column=1, value=acc["account_code"])
            ws.cell(row=row, column=2, value=acc["account_name"])
            cell = ws.cell(row=row, column=3)
            cell.value = float(acc["balance"]) if acc["balance"] != 0 else None
            if acc["balance"] != 0:
                cell.number_format = "#,##0"
            row += 1

        ws.cell(row=row, column=2, value="Tổng doanh thu")
        ws.cell(row=row, column=2).font = Font(bold=True)
        cell = ws.cell(row=row, column=3)
        cell.value = float(data["total_revenue"])
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        row += 2

        ws.cell(row=row, column=1, value="2. GIÁ VỐN HÀNG BÁN")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        for acc in data["expenses"]:
            ws.cell(row=row, column=1, value=acc["account_code"])
            ws.cell(row=row, column=2, value=acc["account_name"])
            cell = ws.cell(row=row, column=3)
            cell.value = float(acc["balance"]) if acc["balance"] != 0 else None
            if acc["balance"] != 0:
                cell.number_format = "#,##0"
            row += 1

        ws.cell(row=row, column=2, value="Tổng giá vốn")
        ws.cell(row=row, column=2).font = Font(bold=True)
        cell = ws.cell(row=row, column=3)
        cell.value = float(data["total_expenses"])
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        row += 2

        ws.cell(row=row, column=2, value="Lợi nhuận gộp")
        ws.cell(row=row, column=2).font = Font(bold=True)
        cell = ws.cell(row=row, column=3)
        cell.value = float(data["gross_profit"])
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        row += 2

        ws.cell(row=row, column=2, value="Lợi nhuận sau thuế")
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        cell = ws.cell(row=row, column=3)
        cell.value = float(data["net_profit_after_tax"])
        cell.number_format = "#,##0"
        cell.font = Font(bold=True, size=12)

        for col in range(1, 4):
            ws.column_dimensions[chr(64 + col)].width = 25

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_journal_voucher(voucher_id: int) -> bytes:
        """Export Journal Voucher to Excel."""
        from services.journal_service import JournalService
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Chung Tu"

        voucher = JournalService.get_voucher(voucher_id)
        if not voucher:
            raise ValueError("Chứng từ không tồn tại")

        title_style = ExcelExporter._create_title_style()

        ws.merge_cells("A1:F1")
        ws["A1"] = "CHỨNG TỪ KẾ TOÁN"
        ws["A1"].font = title_style["font"]
        ws["A1"].alignment = title_style["alignment"]

        ws["A3"] = "Số chứng từ:"
        ws["B3"] = voucher.voucher_no
        ws["A4"] = "Ngày:"
        ws["B4"] = voucher.voucher_date.strftime("%d/%m/%Y")
        ws["A5"] = "Loại:"
        ws["B5"] = voucher.voucher_type
        ws["A6"] = "Diễn giải:"
        ws["B6"] = voucher.description or ""

        headers = ["STT", "Tài khoản", "Mã TK", "Nợ", "Có", "Diễn giải"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=8, column=col, value=header)
            ws.cell(row=8, column=col).font = Font(bold=True)

        for idx, entry in enumerate(voucher.entries, 1):
            ws.cell(row=8 + idx, column=1, value=idx)
            ws.cell(row=8 + idx, column=2, value=entry.account.account_name)
            ws.cell(row=8 + idx, column=3, value=entry.account.account_code)
            ws.cell(row=8 + idx, column=4, value=float(entry.debit) if entry.debit > 0 else None)
            ws.cell(row=8 + idx, column=5, value=float(entry.credit) if entry.credit > 0 else None)
            ws.cell(row=8 + idx, column=6, value=entry.description or "")

        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 18

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
